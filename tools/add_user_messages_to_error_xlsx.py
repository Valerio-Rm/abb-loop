from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook


def make_user_message(static_msg: str | None, orig_msg: str | None) -> str:
    s = (static_msg or "").strip()
    o = (orig_msg or "").strip()
    base = s if s else o
    low = base.lower()

    # Required fields / nulls
    if (
        "cannot be null" in low
        or "values cannot be null" in low
        or "mandatory fields" in low
        or "missing required" in low
        or "required fields" in low
    ):
        return "Compila tutti i campi obbligatori prima di continuare."

    # Date range selection / parameters
    if "p_date_min" in low or "p_date_max" in low or ("date" in low and "null" in low):
        return "Seleziona un intervallo di date valido e riprova."

    # Invalid operations / formats
    if (
        "invalid operation" in low
        or "invalid action" in low
        or "invalid interval" in low
        or "invalid format" in low
        or "use insert or delete" in low
    ):
        return "L'operazione richiesta non è valida. Verifica i dati inseriti e riprova."

    # Not editable / modifiable
    if "not editable" in low or "not modifiable" in low or "cannot update" in low:
        return "Questo elemento non può essere modificato."

    # Not found / invalid selection
    if "not found" in low or "invalid plant_code" in low or "invalid plant_id" in low:
        if "plant" in low:
            return "Lo stabilimento selezionato non è valido o non è disponibile. Verifica la selezione."
        if "line" in low:
            return "La linea selezionata non è valida o non è disponibile. Verifica la selezione."
        if "kpi" in low:
            return "Il KPI selezionato non è valido o non è disponibile. Verifica la selezione."
        if "tier" in low:
            return "Il livello selezionato non è valido o non è disponibile. Verifica la selezione."
        return "L'elemento selezionato non è disponibile. Verifica la selezione e riprova."

    # Timezone/config problems
    if "timezone" in low:
        return (
            "Non è possibile completare l'operazione perché mancano informazioni di configurazione dello stabilimento. "
            "Contatta il supporto."
        )

    # Date validity/overlaps
    if "overlap" in low or "validity interval" in low or "end date" in low or "start date" in low:
        return "Le date inserite non sono valide. Controlla l'intervallo e riprova."

    # Partition/technical backend ops -> generic
    if "partition" in low:
        return "Operazione non disponibile in questo momento. Riprova più tardi."

    # Generic error
    if "error" in low:
        return "Si è verificato un problema durante l'operazione. Riprova; se il problema continua contatta il supporto."

    return "Impossibile completare l'operazione. Verifica i dati e riprova."


def main() -> None:
    xlsx_path = Path(r"c:\Users\Valerio Marignetti\ABB project\list of error message.xlsx")
    out_path = xlsx_path.with_name("list of error message_with_user_message.xlsx")
    wb = load_workbook(xlsx_path)
    ws = wb[wb.sheetnames[0]]

    # Discover headers (row 1)
    header_to_col: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if isinstance(v, str) and v.strip():
            header_to_col[v.strip()] = c

    col_orig = header_to_col.get("error_message_full", 2)
    col_static = header_to_col.get("static_mask", 4)

    # Add/replace output column
    out_header = "user_message"
    if out_header in header_to_col:
        col_out = header_to_col[out_header]
    else:
        col_out = ws.max_column + 1
        ws.cell(1, col_out).value = out_header

    for r in range(2, ws.max_row + 1):
        orig = ws.cell(r, col_orig).value
        static = ws.cell(r, col_static).value
        ws.cell(r, col_out).value = make_user_message(
            static if isinstance(static, str) else None,
            orig if isinstance(orig, str) else None,
        )

    try:
        wb.save(xlsx_path)
    except PermissionError:
        wb.save(out_path)


if __name__ == "__main__":
    main()

