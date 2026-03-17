from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


def is_flagged(v) -> bool:
    if v is None:
        return False
    if v is True:
        return True
    if isinstance(v, (int, float)) and v == 1:
        return True
    if isinstance(v, str) and v.strip().lower() in ("1", "true", "yes", "y", "x"):
        return True
    return False


def better_msg(static_msg: str | None, orig_msg: str | None) -> str:
    s = (static_msg or "").strip()
    o = (orig_msg or "").strip()
    base = s or o
    low = base.lower()

    if "end_date_local must be greater than start_date_local" in low:
        return (
            "Le date inserite non sono valide: la data di fine deve essere successiva alla data di inizio. "
            "Correggi le date e riprova."
        )

    if 'status "to do" not found' in low or "kanban_statuses" in low or ("status" in low and "to do" in low):
        return "Non è possibile completare l'operazione perché manca una configurazione necessaria. Contatta il supporto."

    if "non-working date" in low or "non working" in low or "non_production" in low:
        return "Per salvare la giornata non produttiva devi indicare stabilimento, linea e data. Compila questi campi e riprova."

    if "line does not belong to the specified plant" in low:
        return (
            "La linea selezionata non appartiene allo stabilimento scelto. "
            "Seleziona una linea dello stesso stabilimento e riprova."
        )

    if "line code" in low and "not found" in low:
        return "La linea indicata non è disponibile per lo stabilimento selezionato. Verifica la selezione e riprova."

    if "safety category" in low and "not found" in low:
        return (
            "La categoria di sicurezza indicata non è disponibile per lo stabilimento selezionato. "
            "Verifica la selezione e riprova."
        )

    if "safety type" in low and "not found" in low:
        return "Il tipo di sicurezza indicato non è disponibile. Verifica la selezione e riprova."

    return "Impossibile completare l'operazione con i dati inseriti. Verifica la selezione e riprova."


def main() -> None:
    path = Path(r"c:\Users\Valerio Marignetti\ABB project\list of error message_with_user_message.xlsx")
    wb = load_workbook(path)
    ws = wb[wb.sheetnames[0]]

    # headers
    hdr: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if isinstance(v, str) and v.strip():
            hdr[v.strip()] = c

    col_b = hdr["error_message_full"]
    col_f = hdr["static_mask"]
    col_user = hdr["user_message"]
    col_rev = hdr["to_review"]

    updated = 0
    for r in range(2, ws.max_row + 1):
        if not is_flagged(ws.cell(r, col_rev).value):
            continue
        b = ws.cell(r, col_b).value
        f = ws.cell(r, col_f).value
        ws.cell(r, col_user).value = better_msg(
            f if isinstance(f, str) else None, b if isinstance(b, str) else None
        )
        updated += 1

    wb.save(path)
    print(f"Updated {updated} rows in {path}")


if __name__ == "__main__":
    main()

